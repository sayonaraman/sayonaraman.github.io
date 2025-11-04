"""
Генератор постов для Telegram чата "Мама я вайбкодер"
"""
import os
import re
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from langchain.prompts import ChatPromptTemplate
from langchain.schema import HumanMessage, SystemMessage
from config import (
    OPENROUTER_API_KEY,
    OPENROUTER_MODEL,
    OPENROUTER_API_URL,
    TOPICS_FILE,
    POSTS_DIR,
    POST_STYLE
)


class PostGenerator:
    """Генератор постов через OpenRouter"""
    
    def __init__(self):
        self.api_key = OPENROUTER_API_KEY
        self.model = OPENROUTER_MODEL
        self.api_url = OPENROUTER_API_URL
        
        # Создаем папку для постов
        os.makedirs(POSTS_DIR, exist_ok=True)
        
    def generate_post(self, topic: str) -> str:
        """Генерирует пост по теме через OpenRouter"""
        
        headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json",
            "HTTP-Referer": "https://github.com/MAMA_VIBE",
            "X-Title": "MAMA VIBE Post Generator"
        }
        
        payload = {
            "model": self.model,
            "messages": [
                {
                    "role": "system",
                    "content": POST_STYLE
                },
                {
                    "role": "user",
                    "content": f"Вот мои идеи и мысли для поста:\n\n{topic}\n\nУпакуй это в пост для чата, сохраняя все мои идеи и посылы!"
                }
            ],
            "temperature": 0.8,
            "max_tokens": 1000
        }
        
        try:
            response = requests.post(
                self.api_url,
                headers=headers,
                json=payload,
                timeout=60
            )
            response.raise_for_status()
            
            result = response.json()
            post_content = result['choices'][0]['message']['content']
            
            return post_content.strip()
            
        except Exception as e:
            print(f"❌ Ошибка генерации: {e}")
            return None
    
    def sanitize_filename(self, topic: str) -> str:
        """Создает безопасное имя файла из темы"""
        # Убираем спецсимволы, оставляем буквы, цифры, пробелы
        filename = re.sub(r'[^\w\s-]', '', topic)
        # Заменяем пробелы на подчеркивания
        filename = re.sub(r'\s+', '_', filename)
        # Ограничиваем длину
        filename = filename[:100]
        return filename
    
    def save_post(self, topic: str, content: str) -> str:
        """Сохраняет пост в MD файл"""
        filename = self.sanitize_filename(topic)
        # Сохраняем в папку posts/ для публикации
        filepath = os.path.join("posts", f"{filename}.md")
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(f"# {topic}\n\n")
            f.write(content)
        
        return filepath


class TopicsManager:
    """Управление таблицей с темами"""
    
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.workbook = None
        self.sheet = None
        
    def load(self):
        """Загружает Excel файл"""
        try:
            self.workbook = load_workbook(self.filepath)
            self.sheet = self.workbook.active
            return True
        except FileNotFoundError:
            print(f"❌ Файл {self.filepath} не найден!")
            return False
        except Exception as e:
            print(f"❌ Ошибка загрузки файла: {e}")
            return False
    
    def get_next_topic(self):
        """Находит первую незакрашенную тему"""
        for row_idx, row in enumerate(self.sheet.iter_rows(min_row=2, max_col=1), start=2):
            cell = row[0]
            
            # Проверяем что ячейка не пустая
            if not cell.value:
                continue
            
            # Проверяем что ячейка не закрашена
            if cell.fill.start_color.index == '00000000':  # Нет заливки
                return {
                    'row': row_idx,
                    'topic': cell.value
                }
        
        return None
    
    def mark_as_processed(self, row_idx: int):
        """Закрашивает обработанную строку"""
        cell = self.sheet.cell(row=row_idx, column=1)
        
        # Желтый цвет заливки
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cell.fill = yellow_fill
        
        # Сохраняем изменения
        self.workbook.save(self.filepath)
    
    def close(self):
        """Закрывает файл"""
        if self.workbook:
            self.workbook.close()


def main():
    """Основная функция"""
    print("🚀 Генератор постов для 'Мама я вайбкодер'\n")
    
    # Проверяем API ключ
    if not OPENROUTER_API_KEY:
        print("❌ Не найден OPENROUTER_API_KEY в .env файле!")
        return
    
    # Инициализируем компоненты
    generator = PostGenerator()
    topics = TopicsManager(TOPICS_FILE)
    
    # Загружаем таблицу
    if not topics.load():
        return
    
    # Получаем следующую тему
    next_topic = topics.get_next_topic()
    
    if not next_topic:
        print("✅ Все темы обработаны! Таблица пуста.")
        topics.close()
        return
    
    topic_text = next_topic['topic']
    row_idx = next_topic['row']
    
    print(f"📝 Тема: {topic_text}")
    print(f"⏳ Генерирую пост...\n")
    
    # Генерируем пост
    post_content = generator.generate_post(topic_text)
    
    if not post_content:
        print("❌ Не удалось сгенерировать пост")
        topics.close()
        return
    
    # Сохраняем пост
    filepath = generator.save_post(topic_text, post_content)
    
    # Закрашиваем строку в таблице
    topics.mark_as_processed(row_idx)
    
    print("=" * 60)
    print(post_content)
    print("=" * 60)
    print(f"\n✅ Пост сохранен: {filepath}")
    print(f"✅ Строка {row_idx} закрашена в таблице")
    print(f"\n💡 Чтобы опубликовать пост на сайт, запусти: publish.bat")
    
    topics.close()


if __name__ == "__main__":
    main()

